import json
import openpyxl

with open('Animator_CardStrings (3).json', 'r') as file:
    translate = json.load(file)

with open('Animator_CardStrings (2).json', 'r') as file:
    data = json.load(file)

excel_doc = openpyxl.load_workbook('translation.xlsx')
sheet = excel_doc['Лист1']

counter = 198

count = 2
for i in translate:
    sheet.cell(row=count, column=1).value = i
    sheet.cell(row=count, column=3).value = translate[i]["NAME"]
    sheet.cell(row=count, column=5).value = translate[i]["DESCRIPTION"]
    if translate[i].get("EXTENDED_DESCRIPTION"):
        sheet.cell(row=count, column=7).value = str(translate[i]["EXTENDED_DESCRIPTION"])
    if translate[i].get("UPGRADE_DESCRIPTION"):
        sheet.cell(row=count, column=9).value = translate[i]["UPGRADE_DESCRIPTION"]
    count += 1

count = 2
for i in data:
    sheet.cell(row=count, column=2).value = data[i]["NAME"]
    sheet.cell(row=count, column=4).value = data[i]["DESCRIPTION"]
    if data[i].get("EXTENDED_DESCRIPTION"):
        sheet.cell(row=count, column=6).value = str(data[i]["EXTENDED_DESCRIPTION"])
    if data[i].get("UPGRADE_DESCRIPTION"):
        sheet.cell(row=count, column=8).value = data[i]["UPGRADE_DESCRIPTION"]

    count += 1


excel_doc.save('translation.xlsx')
